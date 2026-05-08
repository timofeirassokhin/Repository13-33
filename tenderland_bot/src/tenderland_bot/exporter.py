"""Excel + Markdown exporters for tender list."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import TenderRow

# Column order — matches the user's requested fields.
COLUMNS: list[tuple[str, str, int]] = [
    # (header, attr, width)
    ("Реестровый номер", "reg_number", 22),
    ("Название", "name", 60),
    ("Начальная цена", "begin_price", 16),
    ("Заказчик", "customer", 40),
    ("Дата публикации", "publish_date", 22),
    ("Дата окончания подачи", "end_date", 22),
    ("Регион", "region", 24),
    ("Тип закупки", "type_name", 28),
    ("Категории лота", "categories", 30),
    ("Источник (ЭТП)", "etp_link", 28),
    ("Документы (zip URL)", "files_url", 50),
    ("Entity ID", "entity_id", 18),
    ("Локальный архив", "local_zip", 40),
]


def _fmt_dt(value: str) -> str:
    """Convert ISO datetime to short Russian-friendly format. Pass through on parse failure."""
    if not value:
        return ""
    try:
        # Tenderland returns "2026-05-05T13:45:55+03:00"
        dt = datetime.fromisoformat(value)
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return value


def write_excel(rows: Iterable[TenderRow], out_path: Path, *, autosearch_name: str) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = (autosearch_name or "tenders")[:31]  # Excel sheet name limit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, _attr, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_header, attr, _width) in enumerate(COLUMNS, start=1):
            value = getattr(row, attr, "")
            if attr in ("publish_date", "end_date"):
                value = _fmt_dt(value)
            elif attr == "begin_price":
                value = float(value or 0.0)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if attr == "begin_price":
                cell.number_format = "# ##0.00 ₽"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def write_markdown(rows: list[TenderRow], out_path: Path, *, autosearch_name: str) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append(f"# Тендеры: {autosearch_name}")
    lines.append("")
    lines.append(f"_Сформировано: {datetime.now().strftime('%Y-%m-%d %H:%M')}_  ")
    lines.append(f"_Всего записей: {len(rows)}_")
    lines.append("")

    for i, row in enumerate(rows, 1):
        lines.append(f"## {i}. {row.reg_number} — {row.name[:120]}")
        lines.append("")
        lines.append(f"- **Начальная цена:** {row.begin_price:,.2f} ₽".replace(",", " "))
        lines.append(f"- **Заказчик:** {row.customer}")
        lines.append(f"- **Регион:** {row.region}")
        lines.append(f"- **Тип:** {row.type_name}")
        lines.append(f"- **Категории:** {row.categories}")
        lines.append(f"- **Опубликован:** {_fmt_dt(row.publish_date)}")
        lines.append(f"- **Дедлайн подачи:** {_fmt_dt(row.end_date)}")
        lines.append(f"- **ЭТП:** {row.etp_link}")
        if row.entity_id:
            lines.append(f"- **Entity ID:** `{row.entity_id}`")
        if getattr(row, "local_zip", ""):
            lines.append(f"- **Архив:** `{row.local_zip}`")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path
