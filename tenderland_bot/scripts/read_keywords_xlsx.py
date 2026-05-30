"""Read the keywords config xlsx and dump readable contents to stdout.

Run: python scripts/read_keywords_xlsx.py <path-to-xlsx> [sheet1 sheet2 ...]
If no sheets passed, dumps all.
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook


def dump_sheet(ws) -> None:
    print(f"\n========== Sheet: {ws.title}  ({ws.max_row}x{ws.max_column}) ==========")
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        if any(cells):
            # tab separator — keeps wide cells aligned-ish in terminal
            print("\t".join(cells))


def main() -> None:
    if len(sys.argv) < 2:
        print("usage: read_keywords_xlsx.py <xlsx> [sheet ...]", file=sys.stderr)
        sys.exit(2)
    path = Path(sys.argv[1])
    only = set(sys.argv[2:]) if len(sys.argv) > 2 else None

    wb = load_workbook(path, data_only=True)
    print(f"File: {path}")
    print(f"Sheets: {wb.sheetnames}")

    for name in wb.sheetnames:
        if only is not None and name not in only:
            continue
        dump_sheet(wb[name])


if __name__ == "__main__":
    main()
