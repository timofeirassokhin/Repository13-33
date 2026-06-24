"""XLSX → text + tables через openpyxl.

Используется для:
  - "ОНМЦК" документов в tender packages (обоснование НМЦК часто в Excel)
  - прайсов производителей
  - наших ТЗ-шаблонов в Excel формате
"""
from __future__ import annotations

import logging
from pathlib import Path

from .core import ExtractedTable, ExtractionOutput

log = logging.getLogger(__name__)


def extract_from_xlsx(path: Path) -> ExtractionOutput:
    out = ExtractionOutput(source_file=path, file_type="xlsx", extractor_name="openpyxl")

    try:
        from openpyxl import load_workbook
    except ImportError:
        out.error = "openpyxl not installed"
        return out

    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        out.error = f"failed to open XLSX: {exc}"
        return out

    text_parts: list[str] = []
    sheet_names = wb.sheetnames
    out.metadata["sheets"] = sheet_names

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)
        if rows:
            out.tables.append(ExtractedTable(rows=rows, title=f"Sheet: {sheet_name}"))
            # Также добавим cell values в общий text
            for row in rows:
                text_parts.append("\t".join(row))
            text_parts.append("")  # пустая строка между листами

    out.text = "\n".join(text_parts).strip()
    out.paragraphs = [p for p in out.text.split("\n") if p.strip()]
    wb.close()

    return out
