"""Analyze a Tenderland export xlsx to assess keyword precision.

Reports volume, customer/supplier/region distribution, price stats, and
heuristic noise classification by keywords in names.
"""
from __future__ import annotations
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


# --- Heuristics for noise detection ----------------------------------------

# Keywords that strongly suggest the tender IS about HPLC/LCMS instrument
SIGNAL_INSTRUMENT = [
    r"хроматограф\w*",
    r"ВЭЖХ", r"УВЭЖХ", r"HPLC", r"UHPLC", r"UPLC",
    r"масс-спектромет\w*", r"масс\s+спектромет\w*",
    r"ЖХ-?МС", r"LC-?MS", r"LCMS",
    r"квадрупол\w*", r"Q-?TOF", r"Orbitrap",
    r"Agilent", r"Shimadzu", r"Waters", r"SCIEX",
    r"Thermo", r"Vanquish", r"Acquity", r"Nexera", r"LCMS-\d+",
]

# Keywords that strongly suggest the tender is NOT what we want
NOISE_PATTERNS = [
    # Reagents / test kits / consumables only (no instrument)
    (r"набор\s*реагент\w*", "reagent_kit"),
    (r"стандартн\w*\s+образ", "standard_sample"),
    (r"^поставк\w+\s+(хроматографическ\w+\s+)?колонок", "columns_only"),
    (r"стандартн\w*\s+образц", "reference_standard"),
    # Lab analytical SERVICES (not instrument)
    (r"оказан\w*\s+услуг\w*\s+(по\s+)?(проведен\w*|выполнен\w*)\s+.*?(анализ|исследован|испытан)", "analysis_service"),
    (r"провед\w+\s+.*?(хроматограф|ВЭЖХ|анализ\w+)", "analysis_service"),
    (r"лабораторн\w*\s+исследован", "lab_research_service"),
    # PCR / molecular (not LC/LCMS)
    (r"\b(ПЦР|qPCR|RT-?PCR)\b", "pcr"),
    (r"\bамплификатор\w*", "pcr"),
    # Other instrument categories that share keywords
    (r"^поставк\w+\s+(газов\w*\s+)?хроматограф", "gas_chromatograph"),  # GC, not LC
    (r"ион\w*\s+хроматограф", "ion_chromatograph"),
    (r"ионообмен\w*\s+хроматограф", "ion_chromatograph"),
    # Calibration / metrology services
    (r"повер\w+\s+(средств\w*\s+)?измерен", "calibration_service"),
    (r"калибровк\w*\s+", "calibration_service"),
    # Repair / maintenance
    (r"ремонт\w*\s+(хроматограф|оборудован|спектромет)", "repair_service"),
    (r"техническ\w*\s+обслуживан", "maintenance"),
    # Software / training
    (r"программн\w*\s+обеспечен", "software"),
    (r"обучен\w+\s+", "training"),
    # Furniture / non-instrument
    (r"мебел\w*", "furniture"),
    (r"вытяжн\w*\s+шкаф", "fume_hood"),
    (r"шкаф\w*\s+вытяжн", "fume_hood"),
    (r"столы\s+лабораторн", "lab_furniture"),
]

CATEGORY_PATTERNS = [
    (r"масс-спектромет|масс\s+спектромет|ЖХ-?МС|LC-?MS|LCMS|квадрупол|Q-?TOF|Orbitrap|Triple\s+Quad", "lc_ms"),
    (r"ВЭЖХ|УВЭЖХ|HPLC|UHPLC|UPLC|жидкостн\w*\s+хроматограф", "hplc"),
    (r"ГПХ|гель-проникающ|GPC|SEC|эксклюзион", "gpc_sec"),
    (r"препаративн\w*", "preparative"),
    (r"газов\w*\s+хроматограф|GC-?MS|GCMS", "gc_or_gcms"),
    (r"ион\w*\s+хроматограф", "ic"),
    (r"спектромет|спектрофотометр", "spectrometry_other"),
]


def classify_row(name: str, subject: str) -> tuple[str, str]:
    """Return (category, noise_label_or_clean)."""
    text = f"{name or ''} {subject or ''}"
    text_l = text.lower()

    # Detect noise first
    for pattern, label in NOISE_PATTERNS:
        if re.search(pattern, text_l, flags=re.IGNORECASE):
            return ("noise", label)

    # Detect category
    for pattern, cat in CATEGORY_PATTERNS:
        if re.search(pattern, text, flags=re.IGNORECASE):
            return (cat, "clean")

    # Has signal but unclassified
    for pattern in SIGNAL_INSTRUMENT:
        if re.search(pattern, text, flags=re.IGNORECASE):
            return ("unclassified_signal", "clean")

    return ("unknown", "no_signal")


def main(path: Path) -> None:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    print(f"Total rows (excl header): {ws.max_row - 1}")
    print(f"Columns: {headers}\n")

    def hidx(name: str, default: int = -1) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return default

    name_idx = hidx("Название")
    subject_idx = hidx("Предмет контракта")
    customer_idx = hidx("Полное наименование заказчика")
    if customer_idx < 0:
        customer_idx = hidx("Наименование заказчика")
    supplier_idx = hidx("Поставщик")
    region_idx = hidx("Регион")
    price_idx = hidx("Начальная цена")
    sum_idx = hidx("Сумма")
    status_idx = hidx("Статус контракта")
    publish_idx = hidx("Дата публикации")
    reg_idx = hidx("Реестровый номер")
    type_idx = hidx("Тип тендера")
    cat_idx = hidx("Категория лота")

    cat_counter: Counter = Counter()
    noise_counter: Counter = Counter()
    customer_counter: Counter = Counter()
    supplier_counter: Counter = Counter()
    region_counter: Counter = Counter()
    status_counter: Counter = Counter()
    by_year: Counter = Counter()
    prices = []
    sums = []

    # Stash some samples per category for later reading
    samples_per_cat: dict[str, list] = {}

    def safe(row, idx):
        return row[idx] if idx >= 0 and idx < len(row) else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if safe(row, name_idx) is None:
            continue
        name = str(safe(row, name_idx) or "")
        subject = str(safe(row, subject_idx) or "")
        customer = str(safe(row, customer_idx) or "")
        supplier = str(safe(row, supplier_idx) or "")
        region = str(safe(row, region_idx) or "")
        status = str(safe(row, status_idx) or "")
        publish = str(safe(row, publish_idx) or "")
        reg = str(safe(row, reg_idx) or "")
        price = safe(row, price_idx)
        contract_sum = safe(row, sum_idx)

        cat, noise = classify_row(name, subject)
        cat_counter[cat] += 1
        if cat == "noise":
            noise_counter[noise] += 1
        else:
            samples_per_cat.setdefault(cat, []).append((reg, name[:120], region, customer[:50]))

        customer_counter[customer] += 1
        supplier_counter[supplier] += 1
        region_counter[region] += 1
        status_counter[status] += 1
        if publish and len(publish) >= 4:
            by_year[publish[:4]] += 1
        if isinstance(price, (int, float)) and price > 0:
            prices.append(float(price))
        if isinstance(contract_sum, (int, float)) and contract_sum > 0:
            sums.append(float(contract_sum))

    print("="*70)
    print("CATEGORY DISTRIBUTION")
    print("="*70)
    total = sum(cat_counter.values())
    for cat, n in sorted(cat_counter.items(), key=lambda x: -x[1]):
        print(f"  {cat:30s} {n:5d}  ({100*n/total:5.1f}%)")

    print()
    print("="*70)
    print("NOISE BREAKDOWN")
    print("="*70)
    for label, n in sorted(noise_counter.items(), key=lambda x: -x[1]):
        print(f"  {label:30s} {n:5d}")

    print()
    print("="*70)
    print("STATUS DISTRIBUTION")
    print("="*70)
    for s, n in sorted(status_counter.items(), key=lambda x: -x[1])[:10]:
        print(f"  {s:50s} {n:5d}")

    print()
    print("="*70)
    print("BY YEAR")
    print("="*70)
    for y, n in sorted(by_year.items()):
        print(f"  {y}  {n:5d}")

    print()
    print("="*70)
    print("PRICE STATS (НМЦК, where > 0)")
    print("="*70)
    if prices:
        prices.sort()
        print(f"  count: {len(prices)}")
        print(f"  median: {prices[len(prices)//2]:>15,.0f} ₽")
        print(f"  mean:   {sum(prices)/len(prices):>15,.0f} ₽")
        print(f"  min:    {prices[0]:>15,.0f} ₽")
        print(f"  max:    {prices[-1]:>15,.0f} ₽")
        print(f"  q25:    {prices[len(prices)//4]:>15,.0f} ₽")
        print(f"  q75:    {prices[3*len(prices)//4]:>15,.0f} ₽")

    print()
    print("="*70)
    print("TOP 15 REGIONS")
    print("="*70)
    for r, n in region_counter.most_common(15):
        print(f"  {r:40s} {n:5d}")

    print()
    print("="*70)
    print("TOP 15 CUSTOMERS")
    print("="*70)
    for c, n in customer_counter.most_common(15):
        print(f"  {n:5d}  {c[:75]}")

    print()
    print("="*70)
    print("TOP 15 SUPPLIERS (winners)")
    print("="*70)
    for s, n in supplier_counter.most_common(15):
        print(f"  {n:5d}  {s[:75]}")

    print()
    print("="*70)
    print("SAMPLES PER CATEGORY (3 each)")
    print("="*70)
    for cat in ["lc_ms", "hplc", "gpc_sec", "preparative", "gc_or_gcms", "ic", "spectrometry_other", "unclassified_signal", "unknown"]:
        if cat in samples_per_cat:
            print(f"\n--- {cat} ---")
            for reg, name, region, customer in samples_per_cat[cat][:3]:
                print(f"  [{reg}] {name}  | {region} | {customer}")


if __name__ == "__main__":
    main(Path(sys.argv[1]))
