# -*- coding: utf-8 -*-
"""Гоним Tier-2 на собранных тендерах + собираем Excel-дайджест в формате пользователя.

Источник данных: _api_check/all/<id>_<topic>.json (собрано scripts/collect_all_samples.py).
ANTHROPIC_API_KEY — в .env (или окружении).

Excel-формат:
  Лист 1 — Дайджест (90%+ блок + 75-90% блок)
  Лист 2 — Отсев (< 75% и blacklist)
  Лист 3 — Tier-3 заглушка (после скачивания файлов)
  Лист 4 — Инфо

Запуск:
    PYTHONPATH=src python scripts/run_tier2_and_excel.py
    PYTHONPATH=src python scripts/run_tier2_and_excel.py --limit-per-topic 10  # только 10 на тему
"""
from __future__ import annotations

import json
import os
import sys
import time
from datetime import date, datetime
from pathlib import Path

# Подгружаем .env (TL_*, ANTHROPIC_API_KEY)
ROOT = Path(__file__).resolve().parents[1]
env_path = ROOT / ".env"
if env_path.exists():
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k, v = k.strip(), v.strip()
        # Перезаписываем пустые значения из системного окружения
        if v and not os.environ.get(k):
            os.environ[k] = v


from tenderland_bot.relevance import Tier2Decision, classify_batch


def parse_args():
    limit = None
    for i, a in enumerate(sys.argv[1:]):
        if a == "--limit-per-topic" and i + 1 < len(sys.argv[1:]):
            limit = int(sys.argv[2 + i])
    return limit


def collect_tasks(limit_per_topic: int | None = None) -> list[tuple[str, str, dict]]:
    """Собрать (tender_id, topic, meta) из _api_check/all/."""
    src = ROOT / "_api_check" / "all"
    out = []
    for jf in sorted(src.glob("*_*.json")):
        if jf.name.startswith("_"):
            continue
        # имя файла: <id>_<topic>.json
        stem = jf.stem
        try:
            sid, topic = stem.split("_", 1)
            int(sid)
        except (ValueError, IndexError):
            continue
        items = json.loads(jf.read_text(encoding="utf-8"))
        if limit_per_topic:
            items = items[:limit_per_topic]
        for it in items:
            t = it.get("tender") or {}
            tid = t.get("regNumber") or it.get("ordinalNumber", "?")
            # Уникализируем — regNumber может совпадать между темами
            out.append((f"{topic}::{tid}", topic, t))
    return out


def build_excel(decisions: list[Tier2Decision], tasks: list, out_path: Path,
                meta_stats: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"
    GREEN_FILL = "C6EFCE"
    YELLOW_FILL = "FFEB9C"
    GREY_FILL = "F2F2F2"
    HDR_FILL = "2E5496"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="FFFFFF", size=14)
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    centerwrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def fill(hex_):
        return PatternFill("solid", fgColor=hex_)

    # Индекс meta по composite key
    meta_by_key = {f"{topic}::{(t.get('regNumber') or '?')}": t for (_, topic, t) in tasks}

    wb = Workbook()
    ws = wb.active
    ws.title = "Дайджест"
    ws.sheet_view.showGridLines = False

    cols = [
        ("№", 5),
        ("Тема", 18),
        ("🔗 Tenderland", 38),
        ("Название", 50),
        ("НМЦК ₽", 16),
        ("Заказчик", 28),
        ("Регион", 16),
        ("Тип закупки", 18),
        ("Подача до", 12),
        ("Поставка до", 12),
        ("Conf", 8),
        ("Класс", 12),
        ("Заказчик-тип", 14),
        ("Сигналы", 35),
        ("Reasoning", 50),
        ("Флаги", 22),
        ("Бренд (T3)", 18),
        ("Наш аналог (T3)", 24),
        ("Папка с файлами", 24),
    ]
    for i, (name, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    c = ws.cell(row=1, column=1, value=f"Tenderland — дайджест от {date.today().isoformat()}")
    c.font = title_font; c.fill = fill(NAVY); c.alignment = centerwrap
    ws.row_dimensions[1].height = 28

    def write_block_header(row, text, color):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
        cc = ws.cell(row=row, column=1, value=text)
        cc.font = bold
        cc.fill = fill(color)
        cc.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22

    def write_columns_header(row):
        for i, (name, _) in enumerate(cols, 1):
            cc = ws.cell(row=row, column=i, value=name)
            cc.font = hdr_font; cc.fill = fill(HDR_FILL); cc.border = border
            cc.alignment = centerwrap

    def fmt_money(v):
        if v is None or v == "":
            return ""
        try:
            return f"{float(v):,.0f}".replace(",", " ")
        except Exception:
            return str(v)

    def fmt_date(v):
        if not v:
            return ""
        try:
            return str(v)[:10]
        except Exception:
            return str(v)

    def tl_url(reg):
        # Tenderland UI URL для тендера
        return f"https://tenderland.ru/Search/Entities?type=Pro&regNumber={reg}" if reg else ""

    def write_row(row_idx, idx_in_block, d: Tier2Decision):
        tid_full = d.tender_id
        topic, reg = tid_full.split("::", 1) if "::" in tid_full else (d.topic, tid_full)
        m = meta_by_key.get(tid_full, {})
        cust = (m.get("customers") or [{}])[0]
        cust_name = cust.get("lotCustomerShortName") or cust.get("customerShortName") or ""

        vals = [
            idx_in_block,
            topic,
            tl_url(reg),
            (m.get("name") or "")[:200],
            fmt_money(m.get("beginPrice")),
            cust_name,
            (m.get("region") or "")[:30],
            (m.get("typeName") or "")[:25],
            fmt_date(m.get("endDate")),
            "",  # Поставка до — заполнится в Tier-3
            f"{d.confidence:.2f}",
            d.detected_class or "",
            d.customer_type or d.customer_class,
            "; ".join(d.matched_signals[:4]),
            d.reasoning[:300],
            "; ".join(d.flags[:4]),
            "",  # Бренд (T3)
            "",  # Наш аналог (T3)
            "",  # Папка
        ]
        for i, v in enumerate(vals, 1):
            cc = ws.cell(row=row_idx, column=i, value=v)
            cc.border = border
            cc.alignment = wrap
            if i == 11:  # confidence
                cc.alignment = centerwrap
            if i == 3 and v:  # URL — гиперссылка
                cc.hyperlink = v
                cc.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 60

    # Сортировка
    high = sorted(
        [d for d in decisions if d.confidence >= 0.90 and d.customer_class != "blacklist"],
        key=lambda x: -x.confidence,
    )
    mid = sorted(
        [d for d in decisions if 0.75 <= d.confidence < 0.90 and d.customer_class != "blacklist"],
        key=lambda x: -x.confidence,
    )
    drop = [d for d in decisions if d.confidence < 0.75 or d.customer_class == "blacklist"]

    r = 3
    write_block_header(r, f"🟢 ВЫСОКАЯ ДОСТОВЕРНОСТЬ (90%+) — {len(high)} тендеров", GREEN_FILL)
    r += 1
    write_columns_header(r); r += 1
    for i, d in enumerate(high, 1):
        write_row(r, i, d); r += 1
    r += 1

    write_block_header(r, f"🟡 СРЕДНЯЯ ДОСТОВЕРНОСТЬ (75-90%) — {len(mid)} тендеров", YELLOW_FILL)
    r += 1
    write_columns_header(r); r += 1
    for i, d in enumerate(mid, 1):
        write_row(r, i, d); r += 1

    ws.freeze_panes = "A2"

    # Лист 2 — Отсев
    ws2 = wb.create_sheet("Отсев")
    ws2.sheet_view.showGridLines = False
    cols2 = [
        ("№", 5),
        ("Тема", 18),
        ("regNumber", 22),
        ("Название", 60),
        ("НМЦК ₽", 14),
        ("Заказчик", 30),
        ("Conf", 8),
        ("Класс заказчика", 14),
        ("Reasoning", 60),
        ("Флаги", 25),
    ]
    for i, (name, w) in enumerate(cols2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for i, (name, _) in enumerate(cols2, 1):
        cc = ws2.cell(row=1, column=i, value=name)
        cc.font = hdr_font; cc.fill = fill(HDR_FILL); cc.border = border; cc.alignment = centerwrap
    rr = 2
    for j, d in enumerate(sorted(drop, key=lambda x: -x.confidence), 1):
        tid_full = d.tender_id
        topic, reg = tid_full.split("::", 1) if "::" in tid_full else (d.topic, tid_full)
        m = meta_by_key.get(tid_full, {})
        cust = (m.get("customers") or [{}])[0]
        vals = [
            j, topic, reg, (m.get("name") or "")[:200],
            f"{float(m.get('beginPrice') or 0):,.0f}".replace(",", " "),
            cust.get("lotCustomerShortName") or "",
            f"{d.confidence:.2f}", d.customer_class,
            d.reasoning[:300], "; ".join(d.flags[:4]),
        ]
        for i, v in enumerate(vals, 1):
            cc = ws2.cell(row=rr, column=i, value=v)
            cc.border = border; cc.alignment = wrap
        ws2.row_dimensions[rr].height = 40
        rr += 1
    ws2.freeze_panes = "A2"

    # Лист 3 — Tier-3 заглушка
    ws3 = wb.create_sheet("Tier-3")
    ws3.merge_cells("A1:H1")
    c3 = ws3.cell(row=1, column=1, value="Tier-3 (скачка ТЗ + анализ) — будет заполняться после следующего этапа")
    c3.font = bold; c3.fill = fill(GREY_FILL); c3.alignment = centerwrap
    ws3.row_dimensions[1].height = 30

    # Лист 4 — Инфо
    ws4 = wb.create_sheet("Инфо")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 35
    info_rows = [
        ("Дата запуска", datetime.now().isoformat(timespec="seconds")),
        ("Всего тендеров на входе", meta_stats["total_in"]),
        ("Из них прошло pass+review (&gt;=75%)", len(high) + len(mid)),
        ("    — высокая &gt;=90%", len(high)),
        ("    — средняя 75-90%", len(mid)),
        ("Отсеяно (< 75% / blacklist)", len(drop)),
        ("Tier-2 модель", meta_stats["model"]),
        ("Прошло секунд", f"{meta_stats['elapsed_sec']:.1f}"),
        ("Input токенов", meta_stats["in_tokens"]),
        ("Output токенов", meta_stats["out_tokens"]),
        ("Cost USD", f"${meta_stats['cost_usd']:.4f}"),
        ("Cost RUB ≈ (80 ₽/$)", f"{meta_stats['cost_usd']*80:.2f} ₽"),
        ("Источник", "_api_check/all/ — собранные 540 тендеров с 21 темы"),
    ]
    for i, (k, v) in enumerate(info_rows, 1):
        c1 = ws4.cell(row=i, column=1, value=k); c1.font = bold
        ws4.cell(row=i, column=2, value=v)

    wb.save(out_path)


def main():
    limit_per_topic = parse_args()
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY не найден ни в .env, ни в окружении.")
        print("Положи его в tenderland_bot/.env как:")
        print("  ANTHROPIC_API_KEY=sk-ant-api...")
        sys.exit(2)

    out_root = ROOT / "_api_check" / "tier2"
    out_root.mkdir(parents=True, exist_ok=True)

    print(">> Сбор задач из _api_check/all/ ...")
    tasks = collect_tasks(limit_per_topic)
    print(f"   найдено {len(tasks)} тендеров для классификации")

    print(">> Гоним Tier-2 (claude-haiku-4-5 + prompt caching) ...")
    t0 = time.time()
    progress_state = {"in": 0, "out": 0, "cost": 0.0}

    def progress(i, total, d):
        progress_state["in"] += d.input_tokens
        progress_state["out"] += d.output_tokens
        progress_state["cost"] += d.cost_usd
        if i % 20 == 0 or i == total:
            print(f"   [{i:>4}/{total}] conf={d.confidence:.2f} {d.relevance:<6} "
                  f"in={progress_state['in']:>6} out={progress_state['out']:>6} "
                  f"cost=${progress_state['cost']:.4f}")

    decisions = classify_batch(tasks, progress_cb=progress)
    elapsed = time.time() - t0
    print(f"   готово за {elapsed:.1f} сек")

    # Сохраним сырые решения
    raw_path = out_root / f"decisions_{date.today().isoformat()}.json"
    raw_path.write_text(
        json.dumps([d.model_dump() for d in decisions], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Excel
    xlsx_path = out_root / f"Tenderland_digest_{date.today().isoformat()}.xlsx"
    meta = {
        "total_in": len(tasks),
        "model": decisions[0].model if decisions else "n/a",
        "elapsed_sec": elapsed,
        "in_tokens": progress_state["in"],
        "out_tokens": progress_state["out"],
        "cost_usd": progress_state["cost"],
    }
    build_excel(decisions, tasks, xlsx_path, meta)

    # Сводка
    high = sum(1 for d in decisions if d.confidence >= 0.90 and d.customer_class != "blacklist")
    mid = sum(1 for d in decisions if 0.75 <= d.confidence < 0.90 and d.customer_class != "blacklist")
    drop = sum(1 for d in decisions if d.confidence < 0.75 or d.customer_class == "blacklist")
    print()
    print(f"  HIGH (&gt;=0.90): {high}")
    print(f"  MID  (0.75-0.90): {mid}")
    print(f"  DROP: {drop}")
    print(f"  cost: ${meta['cost_usd']:.4f}")
    print()
    print(f"Excel → {xlsx_path}")
    print(f"Decisions → {raw_path}")


if __name__ == "__main__":
    main()
