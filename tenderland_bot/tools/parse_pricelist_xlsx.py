"""Universal Glüvex 1С XLSX pricelist parser → JSON.

Парсер унифицированных прайслистов Glüvex (одинаковая структура 14 колонок
для всех поставщиков: MGI, AmoyDX, Burning Rock, Memmert, Hawach, и т.д.).

Структура колонок (case-insensitive, по нормализованным заголовкам):
  Артикул              → vendor_code            (catalog/order number)
  Наименование RU      → description_ru         (для tender-матчинга)
  Наименование EN      → description / model    (для каталога)
  Производитель        → brand
  Валюта               → currency               (Юань/USD/EUR/RUB/etc.)
  Цена закупки         → price_purchase
  Цена продажи РРЦ С НДС / Цена продажи С НДС → price_sale
  НДС                  → vat                    ("без НДС"/"20%"/...)
  Единица              → unit                   (Штука/Упаковка/...)
  Страна происхождения → manufacturer_country
  Поставщик            → distributor / supplier
  ИНН                  → supplier_inn
  Группа               → category_section       (Оборудование/Расходные материалы/...)
  РУ                   → ru_number              (если есть → ru_status='active')

Запуск:
  python tools/parse_pricelist_xlsx.py <path-to-xlsx> [-o output.json] [--brand-override "MGI Tech"]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path


@dataclass
class PriceItem:
    catalog_number: str            # = Артикул (vendor_code)
    description: str               # EN name (Наименование EN)
    description_ru: str = ""       # RU name (для tender матчинга)
    category_section: str = ""     # = Группа
    brand: str = ""                # = Производитель
    currency: str = ""             # = Валюта
    price_purchase: float | None = None
    price_sale: float | None = None
    vat: str = ""
    unit: str = ""
    manufacturer_country: str = ""
    distributor: str = ""
    supplier_inn: str = ""
    ru_number: str = ""            # Из колонки РУ
    is_quote_only: bool = False
    page: int = 0                  # =row_index в XLSX


# ----- Нормализация заголовков → канонический ключ ------------------------

# Канонические ключи. Заголовки в Excel могут быть с пробелами/регистром.
HEADER_ALIASES: dict[str, list[str]] = {
    "catalog_number":         ["артикул", "артикул производителя", "item", "item no", "item number", "article", "sales part no", "part no", "part number", "catalog", "catalog #", "номер артикула", "код товара", "code"],
    "description_ru":         ["наименование на русском", "наименование ru", "наименование товара", "наименование рус", "наименование", "наименование для печати", "название ru", "название на русском", "название"],
    "description":            ["наименование на английском", "наименование en", "наименование eng", "наименование англ", "name en", "наименование (en)", "название на английском", "english name", "рабочее наименование", "sales part no description", "part description", "model", "desciption", "description", "name"],
    "brand":                  ["производитель", "бренд", "manufacturer", "brand"],
    "currency":               ["валюта производителя", "валюта", "currency"],
    "price_purchase":         ["цена закупки с ндс", "цена закупки", "цена закупочная", "цена поставщика", "цена поставщика, евро, без ндс", "transfer price", " transfer  price ", "входящая цена", "входящая цена, chf", "вендат price with vat, rmb", "purchase price"],
    "price_sale":             ["цена продажи/ррц с ндс22%", "цена продажи ррц с ндс", "цена продажи с ндс", "цена продажи", "цена продажи usd", "ррц евро, включая ндс 20%", "ррц евро", "ррц ", "ррц", "sales price", "цена глювекс с ндс, rbm", "цена для клиента, chf", "цена для клиента", "ррц с ндс", "розничная цена", "list price", "sale price"],
    "vat":                    ["ндс", "vat", "tax"],
    "unit":                   ["единица измерения", "единица", "ед. изм.", "unit"],
    "manufacturer_country":   ["страна происхождения", "страна", "country", "origin"],
    "distributor":            ["поставщик (как в 1с)", "поставщик", "supplier", "distributor"],
    "supplier_inn":           ["инн если поставщик рф", "инн", "inn"],
    "category_section":       ["группа", "категория", "group", "category"],
    "ru_number":              ["ру", "ру №", "регудостоверение", "registration"],
}


def _norm(s: str) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower()).replace("ё", "е")


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """Returns mapping canonical_key → column_index_0_based.

    Algorithm (защита от ложных substring-матчей):
      1. Exact normalized match first (для всех ключей).
      2. Затем substring matches — только для алиасов длиной ≥5 chars
         (чтобы "ру" не матчил "русском", "ндс" — "С НДС22%").
      3. Каждый column index может быть назначен только одному ключу.
    """
    norm_headers = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    used_cols: set[int] = set()

    # Pass 1 — exact matches
    for key, aliases in HEADER_ALIASES.items():
        if key in out:
            continue
        for alias in aliases:
            alias_n = _norm(alias)
            for idx, h in enumerate(norm_headers):
                if idx in used_cols:
                    continue
                if h == alias_n:
                    out[key] = idx
                    used_cols.add(idx)
                    break
            if key in out:
                break

    # Pass 2 — startswith matches (для длинных алиасов)
    for key, aliases in HEADER_ALIASES.items():
        if key in out:
            continue
        for alias in aliases:
            alias_n = _norm(alias)
            if len(alias_n) < 5:
                continue
            for idx, h in enumerate(norm_headers):
                if idx in used_cols:
                    continue
                if h.startswith(alias_n):
                    out[key] = idx
                    used_cols.add(idx)
                    break
            if key in out:
                break

    # Pass 3 — substring matches (только для длинных алиасов ≥5 chars)
    for key, aliases in HEADER_ALIASES.items():
        if key in out:
            continue
        for alias in aliases:
            alias_n = _norm(alias)
            if len(alias_n) < 5:
                continue
            for idx, h in enumerate(norm_headers):
                if idx in used_cols:
                    continue
                if alias_n in h:
                    out[key] = idx
                    used_cols.add(idx)
                    break
            if key in out:
                break

    return out


# ----- Парсинг цен --------------------------------------------------------

def _parse_price(cell) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    s = str(cell).strip().replace(" ", "").replace("\xa0", "")
    # Запятая как desimal separator
    s = s.replace(",", ".")
    # Убираем символы валют и буквы
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in (".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ----- Парсинг CSV --------------------------------------------------------

def parse_csv(csv_path: Path) -> tuple[list[PriceItem], dict]:
    """Парсер CSV в том же стиле, что и parse_xlsx.

    Автоматически детектит encoding (utf-8/utf-8-sig/cp1251)
    и delimiter (`,` `;` `\\t`).
    """
    import csv as _csv

    # Detect encoding
    raw = csv_path.read_bytes()
    encoding = "utf-8"
    for enc in ("utf-8-sig", "utf-8", "cp1251", "windows-1251", "latin-1"):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    else:
        return [], {"error": "encoding_detection_failed"}

    # Sniff delimiter
    try:
        dialect = _csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        delimiter = dialect.delimiter
    except _csv.Error:
        delimiter = ","

    rows = list(_csv.reader(text.splitlines(), delimiter=delimiter))
    if not rows:
        return [], {"error": "empty_csv", "encoding": encoding}

    # Find header row
    headers: list[str] = []
    header_row_idx = 0
    for i, row in enumerate(rows[:30], start=1):
        non_empty = sum(1 for c in row if c and c.strip())
        if non_empty >= 3:
            row_lower = " ".join((c or "").lower() for c in row)
            if ("артикул" in row_lower or "наименование" in row_lower
                or "производитель" in row_lower
                or ("article" in row_lower and ("name" in row_lower or "price" in row_lower))
                or ("part no" in row_lower and ("desc" in row_lower or "price" in row_lower or "цена" in row_lower))
                or ("item" in row_lower and ("desc" in row_lower or "price" in row_lower or "цена" in row_lower))):
                headers = [str(c or "").strip() for c in row]
                header_row_idx = i
                break

    if not headers:
        return [], {"error": "header_row_not_found",
                    "encoding": encoding, "delimiter": delimiter,
                    "first_rows": rows[:3]}

    header_map = _build_header_map(headers)
    items: list[PriceItem] = []
    file_stem = csv_path.stem

    for i, row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
        if not row or all(not (c or "").strip() for c in row):
            continue

        def get(key: str) -> str:
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return v.strip() if v else ""

        cat = get("catalog_number")
        if not cat:
            continue
        if cat.lower() in ("итого", "total", "subtotal", "всего"):
            continue
        if len(cat) < 2:
            continue

        desc_en = get("description")
        desc_ru = get("description_ru")
        brand = get("brand") or file_stem.split()[0]
        currency = get("currency")
        vat = get("vat")
        unit = get("unit")
        country = get("manufacturer_country")
        distributor = get("distributor")
        inn = get("supplier_inn")
        section = get("category_section")
        ru_num = get("ru_number")

        purchase_idx = header_map.get("price_purchase")
        sale_idx = header_map.get("price_sale")
        purchase = _parse_price(row[purchase_idx]) if purchase_idx is not None and purchase_idx < len(row) else None
        sale = _parse_price(row[sale_idx]) if sale_idx is not None and sale_idx < len(row) else None

        is_quote = False
        if sale_idx is not None and sale_idx < len(row):
            raw_sale = str(row[sale_idx] or "").lower()
            if "запрос" in raw_sale or "quote" in raw_sale:
                is_quote = True

        items.append(PriceItem(
            catalog_number=cat,
            description=desc_en or desc_ru,
            description_ru=desc_ru,
            brand=brand,
            currency=currency,
            price_purchase=purchase,
            price_sale=sale,
            vat=vat,
            unit=unit,
            manufacturer_country=country,
            distributor=distributor,
            supplier_inn=inn,
            category_section=section,
            ru_number=ru_num,
            is_quote_only=is_quote,
            page=i,
        ))

    # Deduplicate by catalog_number (MEMMERT CSV has many duplicate rows)
    seen: set[str] = set()
    dedup: list[PriceItem] = []
    for it in items:
        if it.catalog_number in seen:
            continue
        seen.add(it.catalog_number)
        dedup.append(it)

    meta = {
        "sheet_name": "csv",
        "encoding": encoding,
        "delimiter": delimiter,
        "header_row_idx": header_row_idx,
        "detected_headers": headers,
        "header_map": header_map,
        "rows_total": len(dedup),
        "rows_before_dedup": len(items),
    }
    return dedup, meta


# ----- Парсинг XLSX -------------------------------------------------------

def parse_xlsx(xlsx_path: Path) -> tuple[list[PriceItem], dict]:
    """Returns (items, meta).

    meta = { detected_headers, header_map, sheet_name, rows_total, ... }
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required: pip install openpyxl", file=sys.stderr)
        raise

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    # Берём первый sheet — гёлвекс всегда так
    ws = wb.active
    sheet_name = ws.title

    rows_iter = ws.iter_rows(values_only=True)

    # Найдём header row — первая строка где >=4 непустых ячеек
    headers: list[str] = []
    header_row_idx = 0
    for i, row in enumerate(rows_iter, start=1):
        non_empty = sum(1 for c in row if c not in (None, ""))
        if non_empty >= 4:
            # Эвристика: содержит "артикул" или "наименование" или "производитель"
            row_lower = " ".join(str(c or "").lower() for c in row)
            if ("артикул" in row_lower or "наименование" in row_lower or "производитель" in row_lower
                or ("item" in row_lower and ("desc" in row_lower or "цена" in row_lower or "price" in row_lower))):
                headers = [str(c or "").strip() for c in row]
                header_row_idx = i
                break

    if not headers:
        return [], {"error": "header_row_not_found", "sheet_name": sheet_name}

    header_map = _build_header_map(headers)
    items: list[PriceItem] = []

    # Default fallback brand из имени файла (пригодится если колонка "Производитель" пуста/одна на всех)
    file_stem = xlsx_path.stem  # e.g. "MGI 2026"

    # Перечисляем остальные строки
    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row:
            continue
        # Skip пустые строки
        if all(c in (None, "") for c in row):
            continue

        def get(key: str) -> str:
            idx = header_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v not in (None, "") else ""

        cat = get("catalog_number")
        if not cat:
            continue
        # Skip "Total", "Итого", "Subtotal" служебные строки
        if cat.lower() in ("итого", "total", "subtotal", "всего"):
            continue
        # Skip заведомо мусорные artikuly (одна буква, длина 1, и т.д.)
        if len(cat) < 2:
            continue

        desc_en = get("description")
        desc_ru = get("description_ru")
        brand = get("brand") or file_stem.split()[0]  # первое слово файла как fallback
        currency = get("currency")
        vat = get("vat")
        unit = get("unit")
        country = get("manufacturer_country")
        distributor = get("distributor")
        inn = get("supplier_inn")
        section = get("category_section")
        ru_num = get("ru_number")

        # Цены — могут быть в любой колонке
        purchase_idx = header_map.get("price_purchase")
        sale_idx = header_map.get("price_sale")
        purchase = _parse_price(row[purchase_idx]) if purchase_idx is not None and purchase_idx < len(row) else None
        sale = _parse_price(row[sale_idx]) if sale_idx is not None and sale_idx < len(row) else None

        # Если sale — строка типа "по запросу" / "Request quote" → is_quote_only
        is_quote = False
        if sale_idx is not None and sale_idx < len(row):
            raw_sale = str(row[sale_idx] or "").lower()
            if "запрос" in raw_sale or "quote" in raw_sale or raw_sale == "по запросу":
                is_quote = True

        items.append(PriceItem(
            catalog_number=cat,
            description=desc_en or desc_ru,
            description_ru=desc_ru,
            brand=brand,
            currency=currency,
            price_purchase=purchase,
            price_sale=sale,
            vat=vat,
            unit=unit,
            manufacturer_country=country,
            distributor=distributor,
            supplier_inn=inn,
            category_section=section,
            ru_number=ru_num,
            is_quote_only=is_quote,
            page=i,
        ))

    meta = {
        "sheet_name": sheet_name,
        "header_row_idx": header_row_idx,
        "detected_headers": headers,
        "header_map": header_map,
        "rows_total": len(items),
    }
    return items, meta


# ----- CLI ----------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(prog="parse_pricelist_xlsx")
    parser.add_argument("xlsx_path", help="Path to Glüvex 1С XLSX pricelist")
    parser.add_argument("-o", "--output", help="Output JSON path (default: <xlsx>.json)")
    parser.add_argument("--brand-override", default="",
                        help="Force brand value (if Производитель column is empty/wrong)")
    parser.add_argument("--distributor-override", default="",
                        help="Force distributor (e.g. 'Glüvex')")
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}", file=sys.stderr)
        return 1

    print(f"Parsing {xlsx_path.name}...")
    if xlsx_path.suffix.lower() == ".csv":
        items, meta = parse_csv(xlsx_path)
    else:
        items, meta = parse_xlsx(xlsx_path)

    if not items:
        print(f"ERROR: no items parsed. meta={meta}", file=sys.stderr)
        return 2

    # Overrides
    if args.brand_override:
        for it in items:
            it.brand = args.brand_override
    if args.distributor_override:
        for it in items:
            it.distributor = args.distributor_override

    print(f"  Sheet: {meta['sheet_name']}")
    print(f"  Header row: {meta['header_row_idx']}")
    print(f"  Detected columns ({len(meta['header_map'])}/{len(HEADER_ALIASES)}):")
    for k, idx in meta["header_map"].items():
        print(f"    {k:24} → col {idx}  ({meta['detected_headers'][idx][:40]})")
    missing = set(HEADER_ALIASES.keys()) - set(meta["header_map"].keys())
    if missing:
        print(f"  Missing columns: {sorted(missing)}")

    print(f"\n  Total items: {len(items)}")

    # Stats
    from collections import Counter
    brands = Counter(it.brand for it in items)
    sections = Counter(it.category_section for it in items if it.category_section)
    countries = Counter(it.manufacturer_country for it in items if it.manufacturer_country)
    with_ru = sum(1 for it in items if it.ru_number)
    with_sale_price = sum(1 for it in items if it.price_sale)

    print(f"  Brands ({len(brands)}):")
    for b, cnt in brands.most_common(10):
        print(f"    {cnt:>5}  {b}")
    print(f"  Groups ({len(sections)}):")
    for s, cnt in sections.most_common(10):
        print(f"    {cnt:>5}  {s[:60]}")
    print(f"  Countries ({len(countries)}):")
    for c, cnt in countries.most_common(5):
        print(f"    {cnt:>5}  {c}")
    print(f"  With РУ number: {with_ru}")
    print(f"  With sale price: {with_sale_price}")

    output = Path(args.output) if args.output else xlsx_path.with_suffix(".json")
    primary_brand = brands.most_common(1)[0][0] if brands else ""

    data = {
        "source_xlsx": str(xlsx_path),
        "brand": primary_brand,
        "items_count": len(items),
        "meta": {
            "sheet_name": meta["sheet_name"],
            "header_row_idx": meta["header_row_idx"],
            "header_columns": {k: meta["detected_headers"][v] for k, v in meta["header_map"].items()},
        },
        "items": [asdict(it) for it in items],
    }
    output.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSaved → {output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
